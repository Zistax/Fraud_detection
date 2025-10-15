import os
from pathlib import Path
import logging
import time
import joblib
import numpy as np
import pandas as pd
import shap
import matplotlib.pyplot as plt
from datetime import datetime
from sklearn.pipeline import Pipeline
from PIL import Image as PILImage
import io
import openpyxl
from typing import Optional
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

logger = logging.getLogger("fraud_batch.shap")
logger.addHandler(logging.NullHandler())

REPORT_DIR = Path("reports")
MODEL_DIR = Path("model")
REPORT_DIR.mkdir(parents=True, exist_ok=True)
MODEL_DIR.mkdir(parents=True, exist_ok=True)


def _strip_resampler_from_pipeline(pipeline):
    try:
        if not hasattr(pipeline, "named_steps"):
            return pipeline
        steps = [(name, step) for name, step in pipeline.named_steps.items() if not hasattr(step, "fit_resample")]
        return Pipeline(steps)
    except Exception:
        return pipeline


def _save_figure_to_png_bytes(fig, dpi=200):
    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=dpi)
    buf.seek(0)
    data = buf.read()
    buf.close()
    return data


def generate_shap_report(pipeline, df, output_dir="reports", batch_name="batch",
                         top_k=1000, sample_frac=0.02, cache_explainer="model/explainer.joblib",
                         chunk_size=2000, max_plot_rows=500):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = int(time.time())
    csv_path = output_dir / f"shap_values_{batch_name}_{ts}.csv"
    png_summary = output_dir / f"shap_summary_{batch_name}_{ts}.png"
    png_bar = output_dir / f"shap_bar_{batch_name}_{ts}.png"

    if df is None or df.empty:
        logger.info("generate_shap_report: df None o vuoto")
        return None, None, None

    # Prepare feature matrix
    X_raw = df.drop(columns=["predicted_fraud", "fraud_proba"], errors="ignore")
    if X_raw.shape[0] == 0:
        return None, None, None

    # Strip resampler
    try:
        infer_pipeline = _strip_resampler_from_pipeline(pipeline)
    except Exception:
        infer_pipeline = pipeline

    # Transform features if preprocess present
    try:
        preprocess = None
        model = infer_pipeline
        if hasattr(infer_pipeline, "named_steps"):
            last_name = list(infer_pipeline.named_steps.keys())[-1]
            model = infer_pipeline.named_steps[last_name]
            preprocess = infer_pipeline.named_steps.get("preproc") or infer_pipeline.named_steps.get("preprocessor")
        if preprocess is not None:
            try:
                X_trans = preprocess.transform(X_raw)
                try:
                    feature_names = preprocess.get_feature_names_out(X_raw.columns)
                except Exception:
                    feature_names = [f"f{i}" for i in range(X_trans.shape[1])]
                X_df = pd.DataFrame(X_trans, columns=feature_names, index=X_raw.index)
            except Exception as e:
                logger.debug("Preprocess transform failed, using raw X. Error: %s", e)
                X_df = X_raw.copy()
        else:
            X_df = X_raw.copy()
    except Exception as e:
        logger.debug("Feature transform error, using raw. %s", e)
        X_df = X_raw.copy()
        model = infer_pipeline

    n = len(X_df)
    if n == 0:
        return None, None, None

    # Select rows by probability or sample
    try:
        probs = None
        if hasattr(infer_pipeline, "predict_proba"):
            probs = infer_pipeline.predict_proba(X_raw)[:, 1]
        elif hasattr(model, "predict_proba"):
            probs = model.predict_proba(X_df)[:, 1]

        if probs is not None:
            idx = np.argsort(probs)[-min(top_k, n):]
            X_sel = X_df.iloc[idx]
        else:
            take = min(n, int(max(1, n * sample_frac)))
            X_sel = X_df.sample(take, random_state=42)
    except Exception as e:
        logger.debug("Error selecting rows by probability: %s. Falling back to sampling.", e)
        X_sel = X_df.sample(min(n, top_k), random_state=42)

    # Load or create explainer
    cache_expl = Path(cache_explainer)
    explainer = None
    if cache_expl.exists():
        try:
            explainer = joblib.load(str(cache_expl))
        except Exception:
            explainer = None

    if explainer is None:
        try:
            sample_data = shap.sample(X_df, min(100, len(X_df)))
            try:
                explainer = shap.TreeExplainer(model, data=sample_data)
            except Exception:
                explainer = shap.Explainer(model, masker=shap.maskers.Independent(sample_data))
            try:
                cache_expl.parent.mkdir(parents=True, exist_ok=True)
                joblib.dump(explainer, str(cache_expl))
            except Exception:
                logger.debug("Non è stato possibile salvare la cache dell'explainer.")
        except Exception as e:
            logger.exception("Errore creazione explainer SHAP: %s", e)
            return None, None, None

    # Compute SHAP values in chunks, save CSV, accumulate subset for plotting
    try:
        total_rows = len(X_sel)
        vals_list = []
        X_for_plot = []
        write_header = True
        rows_acc = 0

        for start in range(0, total_rows, chunk_size):
            X_chunk = X_sel.iloc[start:start+chunk_size]
            sv = explainer(X_chunk)
            vals = getattr(sv, "values", None)
            if vals is None:
                vals = np.array(sv)
            df_vals = pd.DataFrame(vals, columns=X_chunk.columns, index=X_chunk.index)
            mode = "w" if write_header else "a"
            df_vals.to_csv(csv_path, mode=mode, header=write_header)
            write_header = False

            if rows_acc < max_plot_rows:
                need = max_plot_rows - rows_acc
                take = min(len(X_chunk), need)
                if take > 0:
                    vals_list.append(vals[:take])
                    X_for_plot.append(X_chunk.iloc[:take])
                    rows_acc += take

        if vals_list:
            vals_plot = np.vstack(vals_list)
            X_plot = pd.concat(X_for_plot, axis=0).iloc[:len(vals_plot)]

            # --- summary plot ---
            try:
                fig, ax = plt.subplots(figsize=(10, 6))
                shap.summary_plot(vals_plot, X_plot, max_display=30, show=False, plot_type="dot", color=plt.get_cmap("coolwarm"))
                plt.tight_layout()
                png_bytes = _save_figure_to_png_bytes(fig, dpi=250)
                plt.close(fig)
                with open(png_summary, "wb") as f:
                    f.write(png_bytes)
            except Exception as e:
                logger.warning("Impossibile creare summary plot PNG: %s", e)
                png_summary = None

            # --- bar plot ---
            try:
                mean_abs = np.abs(vals_plot).mean(axis=0)
                feat_imp = pd.Series(mean_abs, index=X_plot.columns).sort_values(ascending=False).head(30)
                fig2, ax2 = plt.subplots(figsize=(8, max(4, len(feat_imp)*0.25)))
                feat_imp.sort_values().plot(kind="barh", ax=ax2)
                ax2.set_xlabel("Mean |SHAP|")
                ax2.set_title("Feature importance (mean |SHAP|)")
                fig2.tight_layout()
                png_bar_bytes = _save_figure_to_png_bytes(fig2, dpi=250)
                plt.close(fig2)
                with open(png_bar, "wb") as f:
                    f.write(png_bar_bytes)
            except Exception as e:
                logger.warning("Impossibile creare bar plot PNG: %s", e)
                png_bar = None

        else:
            png_summary = None
            png_bar = None

    except Exception as e:
        logger.exception("Errore calcolo SHAP: %s", e)
        return None, None, None

    return (str(png_summary) if png_summary and Path(png_summary).exists() else None,
            str(png_bar) if png_bar and Path(png_bar).exists() else None,
            str(csv_path) if csv_path.exists() else None)


def save_detected_frauds_to_excel(
    df: pd.DataFrame,
    shap_csv_path: Optional[str] = None,
    batch_name: str = "batch",
    output_dir: str = "reports",
    max_rows: int = 50,
    round_decimals: int = 3
) -> Optional[str]:
    try:
        outdir = Path(output_dir)
        outdir.mkdir(parents=True, exist_ok=True)
        ts = int(time.time())
        excel_path = outdir / f"detected_frauds_{batch_name}_{ts}.xlsx"

        # validate input
        if df is None or df.empty or 'predicted_fraud' not in df.columns:
            return None

        frauds = df[df['predicted_fraud'] == 1].copy()
        if frauds.empty:
            return None

        # preview limit
        frauds = frauds.head(max_rows)

        # --- column ordering: prefer informative cols first ---
        preferred_order = ["Time", "time", "transaction_time", "TransactionID", "id",
                           "Amount", "amount", "Class", "class", "fraud_proba", "predicted_fraud"]
        cols = list(frauds.columns)
        ordered = []
        # add first matching preferred cols keeping their canonical form if present
        for pref in preferred_order:
            if pref in frauds.columns and pref not in ordered:
                ordered.append(pref)
        # then add remaining columns in original order
        for c in cols:
            if c not in ordered:
                ordered.append(c)
        frauds = frauds.loc[:, ordered]

        # round numeric columns for presentation
        num_cols = frauds.select_dtypes(include=[np.number]).columns.tolist()
        for c in num_cols:
            frauds[c] = frauds[c].round(round_decimals)

        # write to excel using pandas then style with openpyxl
        frauds.to_excel(excel_path, sheet_name="Detected_Frauds", index=False)

        # open workbook for styling
        wb = openpyxl.load_workbook(excel_path)
        ws = wb["Detected_Frauds"]

        # header style: bold and centered vertically
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", horizontal="center")

        # freeze header row
        ws.freeze_panes = ws["A2"]

        # autofilter full table
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row >= 1 and max_col >= 1:
            last_col_letter = get_column_letter(max_col)
            ws.auto_filter.ref = f"A1:{last_col_letter}{max_row}"

        # auto column widths (based on content)
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx, max_row=ws.max_row):
                for cell in row:
                    try:
                        value = "" if cell.value is None else str(cell.value)
                    except Exception:
                        value = str(cell.value)
                    if len(value) > max_length:
                        max_length = len(value)
            adjusted_width = (max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        # Number formats: detect specific columns
        col_name_to_idx = {ws.cell(row=1, column=i).value: i for i in range(1, max_col + 1)}

        # Amount format (if present)
        amount_candidates = [c for c in ("Amount", "amount") if c in col_name_to_idx]
        if amount_candidates:
            amt_col = col_name_to_idx[amount_candidates[0]]
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=amt_col)
                # apply a currency-like format or two decimals
                cell.number_format = '#,##0.00'

        # fraud_proba format and conditional color scale
        if "fraud_proba" in col_name_to_idx:
            prob_col = col_name_to_idx["fraud_proba"]
            prob_col_letter = get_column_letter(prob_col)
            # set number format
            for r in range(2, ws.max_row + 1):
                ws.cell(row=r, column=prob_col).number_format = '0.000'
            # apply color scale: green (low) -> white (mid) -> red (high)
            cs_rule = ColorScaleRule(start_type='min', start_color='63BE7B',
                                     mid_type='percentile', mid_value=50, mid_color='FFFFFF',
                                     end_type='max', end_color='FF6B6B')
            ws.conditional_formatting.add(f"{prob_col_letter}2:{prob_col_letter}{ws.max_row}", cs_rule)

        # predicted_fraud highlight (cells equal 1)
        if "predicted_fraud" in col_name_to_idx:
            pf_col = col_name_to_idx["predicted_fraud"]
            pf_col_letter = get_column_letter(pf_col)
            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
            rule = CellIsRule(operator='equal', formula=['1'], stopIfTrue=True, fill=fill)
            ws.conditional_formatting.add(f"{pf_col_letter}2:{pf_col_letter}{ws.max_row}", rule)

        # optionally: highlight entire row when predicted_fraud == 1 (light red tint)
        # (This is more intrusive and might slow down very large sheets; enable if desired)
        try:
            if "predicted_fraud" in col_name_to_idx:
                pf_col = col_name_to_idx["predicted_fraud"]
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(row=r, column=pf_col)
                    try:
                        cell_val = cell.value
                        # treat both int 1 and string '1'
                        if cell_val == 1 or str(cell_val).strip() == "1":
                            # apply subtle fill to whole row
                            row_fill = PatternFill(start_color="FFF7F7", end_color="FFF7F7", fill_type="solid")
                            for c_idx in range(1, max_col + 1):
                                ws.cell(row=r, column=c_idx).fill = row_fill
                    except Exception:
                        continue
        except Exception:
            # do not fail entire process on styling problems
            pass

        # save workbook
        wb.save(excel_path)
        return str(excel_path)

    except Exception as e:
        # if something goes wrong, log (if you have logger) or print and return None
        try:
            import logging
            logging.getLogger("fraud_batch.shap").exception("Errore salvataggio frodi in Excel: %s", e)
        except Exception:
            print("Errore salvataggio frodi in Excel:", e)
        return None